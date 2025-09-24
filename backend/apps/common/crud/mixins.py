"""CRUD Framework Mixins."""

from __future__ import annotations

from typing import TYPE_CHECKING, Any, Protocol

if TYPE_CHECKING:
    from django.db.models import Model, QuerySet
    from django.http import HttpRequest, HttpResponse

import csv
from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.core.exceptions import ImproperlyConfigured
from django.db.models import Q, QuerySet
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .config import CRUDConfig, FieldConfig
from .utils import format_field_value, get_field_value


class ViewProtocol(Protocol):
    """Protocol for Django view-like objects."""

    request: HttpRequest
    model: type[Model] | None

    def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse: ...
    def get_queryset(self) -> QuerySet[Any]: ...
    def get_context_data(self, **kwargs: Any) -> dict[str, Any]: ...
    def render_to_response(self, context: dict[str, Any], **response_kwargs: Any) -> HttpResponse: ...


class FormViewProtocol(ViewProtocol, Protocol):
    """Protocol for form view-like objects."""

    def get_success_url(self) -> str: ...
    def form_valid(self, form: Any) -> HttpResponse: ...


class DetailViewProtocol(ViewProtocol, Protocol):
    """Protocol for detail view-like objects."""

    object: Model | None

    def get_object(self, queryset: QuerySet[Any] | None = None) -> Model: ...


class CRUDConfigMixin:
    """Base mixin that provides CRUD configuration."""

    crud_config: CRUDConfig | None = None

    def get_crud_config(self) -> CRUDConfig:
        """Get the CRUD configuration."""
        if self.crud_config is None:
            raise ImproperlyConfigured(f"{self.__class__.__name__} must define crud_config attribute")
        return self.crud_config

    def get_field_configs(self) -> list[FieldConfig]:
        """Get field configurations."""
        config = self.get_crud_config()
        if not config.fields:
            # Auto-generate from model if not specified
            return self._generate_field_configs()
        return config.fields

    def _generate_field_configs(self) -> list[FieldConfig]:
        """Auto-generate field configs from model."""
        if not hasattr(self, "model") or not self.model:
            return []

        fields = []
        for field in self.model._meta.fields:
            if field.name in ["password", "secret"]:  # Skip sensitive fields
                continue

            field_type = "text"
            if field.get_internal_type() in [
                "IntegerField",
                "DecimalField",
                "FloatField",
            ]:
                field_type = "number"
            elif field.get_internal_type() == "BooleanField":
                field_type = "boolean"
            elif field.get_internal_type() in ["DateField", "DateTimeField"]:
                field_type = field.get_internal_type().replace("Field", "").lower()
            elif field.get_internal_type() == "ForeignKey":
                field_type = "foreign_key"
            elif field.get_internal_type() == "ImageField":
                field_type = "image"

            fields.append(
                FieldConfig(
                    name=field.name,
                    verbose_name=field.verbose_name,
                    field_type=field_type,
                    searchable=field_type == "text",
                ),
            )

        return fields


class CRUDListMixin(CRUDConfigMixin, PermissionRequiredMixin):
    """Mixin for list views with search, sort, filter, and export."""

    # Type annotations for Django view attributes
    request: HttpRequest
    model: type[Model] | None = None

    def get_permission_required(self) -> list[str]:
        """Get required permissions."""
        config = self.get_crud_config()
        if config.list_permission:
            return [config.list_permission]
        # Default to Django's view permission
        if hasattr(self, "model") and self.model:
            return [f"{self.model._meta.app_label}.view_{self.model._meta.model_name}"]
        return []

    def get_template_names(self) -> list[str]:
        """Get template name, supporting HTMX partials."""
        config = self.get_crud_config()
        if self.request.headers.get("HX-Request"):
            return [config.table_template]
        return [config.list_template]

    def get_paginate_by(self, queryset: QuerySet[Any]) -> int:
        """Get pagination size from request or config."""
        config = self.get_crud_config()
        try:
            page_size = int(self.request.GET.get("page_size", config.paginate_by))
            if page_size in config.paginate_choices:
                return page_size
        except (ValueError, TypeError):
            pass
        return config.paginate_by

    def get_queryset(self) -> QuerySet[Any]:
        """Get queryset with search and sort."""
        queryset = super().get_queryset()  # type: ignore[misc]

        # Apply search with input validation
        search_query = self.request.GET.get("search", "").strip()
        if search_query and self._is_valid_search_query(search_query):
            queryset = self.apply_search(queryset, search_query)

        # Apply sorting with validation to prevent injection
        sort_by = self.request.GET.get("sort_by", "")
        if sort_by and self._is_valid_sort_field(sort_by):
            direction = self.request.GET.get("direction", "asc")
            if direction == "desc":
                sort_by = f"-{sort_by}"
            queryset = queryset.order_by(sort_by)
        else:
            config = self.get_crud_config()
            queryset = queryset.order_by(config.default_sort_field)

        return queryset

    def _is_valid_sort_field(self, field_name: str) -> bool:
        """Validate that sort field is allowed to prevent injection attacks."""
        # Remove minus sign for desc sorting
        clean_field = field_name.lstrip("-")

        # Get valid field names from model and field configs
        valid_fields = set()

        # Add model field names
        if hasattr(self, "model") and self.model:
            for field in self.model._meta.fields:
                valid_fields.add(field.name)

            # Add foreign key relations (one level deep)
            for field in self.model._meta.fields:
                if field.get_internal_type() == "ForeignKey":
                    related_model = field.related_model
                    if related_model and hasattr(related_model, "_meta") and not isinstance(related_model, str):
                        for related_field in related_model._meta.fields:
                            valid_fields.add(f"{field.name}__{related_field.name}")

        # Add field config names
        for field_config in self.get_field_configs():
            if field_config.sortable:
                valid_fields.add(field_config.name)

        return clean_field in valid_fields

    def _is_valid_search_query(self, search_query: str) -> bool:
        """Validate search query to prevent malicious input."""
        # Limit search query length to prevent DoS
        if len(search_query) > 255:
            return False

        # Prevent SQL injection attempts - block dangerous patterns
        dangerous_patterns = [
            r"--",  # SQL comments
            r"/\*",  # SQL block comments start
            r"\*/",  # SQL block comments end
            r";",  # SQL statement separator
            r"\bUNION\b",  # SQL UNION
            r"\bSELECT\b",  # SQL SELECT
            r"\bINSERT\b",  # SQL INSERT
            r"\bUPDATE\b",  # SQL UPDATE
            r"\bDELETE\b",  # SQL DELETE
            r"\bDROP\b",  # SQL DROP
            r"\bEXEC\b",  # SQL EXEC
            r"<script",  # XSS script tags
            r"javascript:",  # JavaScript URLs
        ]

        import re

        for pattern in dangerous_patterns:
            if re.search(pattern, search_query, re.IGNORECASE):
                return False

        return True

    def apply_search(self, queryset: QuerySet[Any], search_query: str) -> QuerySet[Any]:
        """Apply search to queryset."""
        search_fields = self.get_search_fields()
        if not search_fields:
            return queryset

        # Validate search fields against model to prevent field injection
        validated_fields = self._validate_search_fields(search_fields)
        if not validated_fields:
            return queryset

        q_objects = Q()
        for field in validated_fields:
            q_objects |= Q(**{f"{field}__icontains": search_query})

        return queryset.filter(q_objects)

    def get_search_fields(self) -> list[str]:
        """Get searchable fields."""
        fields = []
        for field_config in self.get_field_configs():
            if field_config.searchable:
                fields.append(field_config.name)
        return fields

    def _validate_search_fields(self, search_fields: list[str]) -> list[str]:
        """Validate search fields against model to prevent field injection attacks."""
        if not hasattr(self, "model") or not self.model:
            return search_fields

        valid_fields = set()

        # Add model field names (direct fields only, no lookups)
        for field in self.model._meta.fields:
            valid_fields.add(field.name)

        # Add many-to-many field names
        for field in self.model._meta.many_to_many:
            valid_fields.add(field.name)

        # Validate each search field
        validated_fields = []
        for field_name in search_fields:
            # Check if field exists in model
            if field_name in valid_fields:
                validated_fields.append(field_name)

        return validated_fields

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Add CRUD context."""
        context = super().get_context_data(**kwargs)
        config = self.get_crud_config()

        field_configs = self.get_field_configs()
        context.update(
            {
                "crud": config,  # Template expects 'crud' not 'crud_config'
                "crud_config": config,  # Keep for backwards compatibility
                "field_configs": field_configs,
                "table_fields": field_configs,
                "page_title": config.page_title,
                "page_subtitle": config.page_subtitle,
                "page_icon": config.page_icon,
                "create_url": self.get_create_url(),
                "add_url": self.get_create_url(),
                "export_formats": config.export_formats if config.enable_export else [],
                "paginate_choices": config.paginate_choices,
                "current_page_size": self.get_paginate_by(self.get_queryset()),
            },
        )

        # Add extra context
        context.update(config.extra_context)

        # Add request params for maintaining state
        context["request_params"] = self.request.GET.copy()

        return context

    def get_create_url(self) -> str | None:
        """Get create URL."""
        config = self.get_crud_config()
        if config.create_url_name:
            return reverse(config.create_url_name)
        return None

    def render_to_response(self, context: dict[str, Any], **response_kwargs: Any) -> HttpResponse:
        """Handle export requests."""
        export_format = self.request.GET.get("format")
        if export_format in ["csv", "xlsx"]:
            return self.export_data(export_format, context["object_list"])

        return super().render_to_response(context, **response_kwargs)

    def export_data(self, format_type: str, queryset: QuerySet[Any]) -> HttpResponse:
        """Export data to CSV or XLSX with size limits and memory protection."""
        config = self.get_crud_config()

        # Check export size limits to prevent memory exhaustion
        MAX_EXPORT_SIZE = 10000  # Maximum records per export
        queryset_count = queryset.count()

        if queryset_count > MAX_EXPORT_SIZE:
            messages.error(
                self.request,
                f"Export size ({queryset_count:,} records) exceeds maximum limit "
                f"({MAX_EXPORT_SIZE:,} records). Please add filters to reduce the dataset.",
            )
            return self.get(self.request)  # type: ignore[attr-defined]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.export_filename_prefix}_{timestamp}.{format_type}"

        if format_type == "csv":
            return self.export_csv(queryset, filename)
        elif format_type == "xlsx" and HAS_OPENPYXL:
            return self.export_xlsx(queryset, filename)
        else:
            messages.error(self.request, f"Export format '{format_type}' not supported")
            return self.get(self.request)  # type: ignore[attr-defined]

    def export_csv(self, queryset: QuerySet[Any], filename: str) -> HttpResponse:
        """Export to CSV with memory-efficient streaming."""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        field_configs = [f for f in self.get_field_configs() if f.export]

        # Write headers
        headers = [f.verbose_name for f in field_configs]
        writer.writerow(headers)

        # Write data in chunks to prevent memory exhaustion
        CHUNK_SIZE = 1000
        for chunk in queryset.iterator(chunk_size=CHUNK_SIZE):
            row = []
            for field_config in field_configs:
                try:
                    value = get_field_value(chunk, field_config.name)
                    formatted_value = format_field_value(value, field_config)
                    row.append(formatted_value)
                except (AttributeError, ValueError, TypeError):
                    # Handle cases where field doesn't exist or can't be formatted
                    row.append("")
            writer.writerow(row)

        return response

    def export_xlsx(self, queryset: QuerySet[Any], filename: str) -> HttpResponse:
        """Export to XLSX."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is None:
            worksheet = workbook.create_sheet("Data Export")
        worksheet.title = "Data Export"

        field_configs = [f for f in self.get_field_configs() if f.export]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write headers
        for col, field_config in enumerate(field_configs, 1):
            cell = worksheet.cell(row=1, column=col, value=field_config.verbose_name)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, field_config in enumerate(field_configs, 1):
                value = get_field_value(obj, field_config.name)
                formatted_value = format_field_value(value, field_config)
                worksheet.cell(row=row_idx, column=col_idx, value=formatted_value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            # Get column letter safely - skip if first cell is a MergedCell
            first_cell = column[0]
            if not hasattr(first_cell, "column_letter"):
                continue
            column_letter = first_cell.column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, ValueError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        virtual_workbook = BytesIO()
        workbook.save(virtual_workbook)
        response.write(virtual_workbook.getvalue())

        return response


class CRUDFormMixin(CRUDConfigMixin, PermissionRequiredMixin):
    """Mixin for create/update views."""

    # Subclasses are expected to provide a Django model on the class
    model: Any | None = None

    def get_template_names(self) -> list[str]:
        """Get template name."""
        config = self.get_crud_config()
        return [config.form_template]

    def get_success_url(self) -> str:
        """Get success URL."""
        config = self.get_crud_config()
        if config.list_url_name:
            return reverse_lazy(config.list_url_name)
        return super().get_success_url()

    def form_valid(self, form: Any) -> HttpResponse:
        """Add success message."""
        response = super().form_valid(form)
        action = "created" if isinstance(self, CreateView) else "updated"
        messages.success(
            self.request,  # type: ignore[attr-defined]
            f"{self.model._meta.verbose_name.title()} {action} successfully.",  # type: ignore[attr-defined]
        )
        return response

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Add CRUD context."""
        context = super().get_context_data(**kwargs)
        config = self.get_crud_config()

        context.update(
            {
                "crud_config": config,
                "page_title": self.get_page_title(),
                "page_subtitle": config.page_subtitle,
                "page_icon": config.page_icon,
                "cancel_url": self.get_cancel_url(),
            },
        )

        context.update(config.extra_context)
        return context

    def get_page_title(self) -> str:
        """Get page title."""
        if isinstance(self, CreateView):
            return f"Add {self.model._meta.verbose_name.title()}"
        else:
            return f"Edit {self.model._meta.verbose_name.title()}"

    def get_cancel_url(self) -> str:
        """Get cancel URL."""
        config = self.get_crud_config()
        if config.list_url_name:
            return reverse(config.list_url_name)
        return "/"


class CRUDDetailMixin(CRUDConfigMixin, PermissionRequiredMixin):
    """Mixin for detail views."""

    def get_permission_required(self) -> list[str]:
        """Get required permissions."""
        config = self.get_crud_config()
        if config.list_permission:
            return [config.list_permission]
        if hasattr(self, "model") and self.model:
            return [f"{self.model._meta.app_label}.view_{self.model._meta.model_name}"]
        return []

    def get_template_names(self) -> list[str]:
        """Get template name."""
        config = self.get_crud_config()
        return [config.detail_template]

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Add CRUD context."""
        context = super().get_context_data(**kwargs)
        config = self.get_crud_config()

        context.update(
            {
                "crud_config": config,
                "field_configs": self.get_field_configs(),
                "page_title": str(self.object),  # type: ignore[attr-defined]
                "page_subtitle": config.page_subtitle,
                "page_icon": config.page_icon,
                "edit_url": self.get_edit_url(),
                "delete_url": self.get_delete_url(),
                "list_url": self.get_list_url(),
            },
        )

        context.update(config.extra_context)
        return context

    def get_edit_url(self) -> str | None:
        """Get edit URL."""
        config = self.get_crud_config()
        if config.update_url_name:
            return reverse(config.update_url_name, kwargs={"pk": self.object.pk})  # type: ignore[attr-defined]
        return None

    def get_delete_url(self) -> str | None:
        """Get delete URL."""
        config = self.get_crud_config()
        if config.delete_url_name:
            return reverse(config.delete_url_name, kwargs={"pk": self.object.pk})  # type: ignore[attr-defined]
        return None

    def get_list_url(self) -> str | None:
        """Get list URL."""
        config = self.get_crud_config()
        if config.list_url_name:
            return reverse(config.list_url_name)
        return None


class CRUDDeleteMixin(CRUDConfigMixin, PermissionRequiredMixin):
    """Mixin for delete views."""

    def get_permission_required(self) -> list[str]:
        """Get required permissions."""
        config = self.get_crud_config()
        if config.delete_permission:
            return [config.delete_permission]
        if hasattr(self, "model") and self.model:
            return [f"{self.model._meta.app_label}.delete_{self.model._meta.model_name}"]
        return []

    def get_template_names(self) -> list[str]:
        """Get template name."""
        config = self.get_crud_config()
        return [config.delete_template]

    def get_success_url(self) -> str:
        """Get success URL."""
        config = self.get_crud_config()
        if config.list_url_name:
            return reverse_lazy(config.list_url_name)
        return super().get_success_url()

    def delete(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        """Add success message."""
        self.object = self.get_object()  # type: ignore[attr-defined]
        success_url = self.get_success_url()
        self.object.delete()
        messages.success(request, f"{self.model._meta.verbose_name.title()} deleted successfully.")  # type: ignore[attr-defined]
        return HttpResponseRedirect(success_url)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Add CRUD context."""
        context = super().get_context_data(**kwargs)
        config = self.get_crud_config()

        context.update(
            {
                "crud_config": config,
                "page_title": f"Delete {self.model._meta.verbose_name.title()}",  # type: ignore[attr-defined]
                "page_subtitle": "This action cannot be undone.",
                "page_icon": config.page_icon,
                "cancel_url": self.get_cancel_url(),
            },
        )

        context.update(config.extra_context)
        return context

    def get_cancel_url(self) -> str:
        """Get cancel URL."""
        config = self.get_crud_config()
        if config.list_url_name:
            return reverse(config.list_url_name)
        return "/"
